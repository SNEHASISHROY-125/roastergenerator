import openpyxl as opxl , random , os
from openpyxl.styles import Font , PatternFill , Border, Side , Alignment


# Create a workbook
wb = opxl.Workbook()

# Select the active worksheet
ws = wb.active

for i in range(1,10):
    # ws.append([])  
    font = Font(size=14,bold=True,italic=True)
    _ = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
    ws.cell(row=1, column=i, value=random.choice(_))  # Write 'World' to cell B1
    ws.cell(row=1, column=i).font = font
     
# Write values to the cells of a row-1
# ws.cell(row=1, column=1, value='01')  # Write 'Hello' to cell A1
# Write values to the cells of a row-2
[ws.cell(row=2, column=i, value='Hello') for i in range(1,10)]  # Write 'Hello' to cell A1

# Adjust column widths
for column_cells in ws.columns:
    length = 20
    ws.column_dimensions[column_cells[0].column_letter].width = length

# set the font style-border-fill
    # Create a Side object with desired style
side = Side(border_style='thin', color='000000')  # '000000' is black

# Create a Border object using the Side object
border = Border(top=side, right=side, bottom=side, left=side)

# Apply the border to the cells
for row in ws.iter_rows():
    for cell in row:
        cell.border = border

# Set the fill color of the cells
fill = PatternFill(fill_type='solid', fgColor='FFFF00')  # Yellow fill
for row in ws.iter_rows(min_row=2, max_row=2):
    for cell in row:
        cell.fill = fill

# Incert a new row
ws.insert_rows(1)
# Set the alignment of the text in the cells
c=ws.cell(row=1, column=1, value='WEEKLY ROASTER MAR 2024').font = Font(size=26, bold=True, italic=True)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A1:I1')

# Save the workbook
wb.save('sample.xlsx')

# Open the workbook
os.system('start excel sample.xlsx')