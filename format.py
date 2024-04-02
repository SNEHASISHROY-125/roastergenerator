'''
This script generates a formatted Excel workbook for a weekly roaster.

The script uses the openpyxl library to create a workbook and format its contents. It writes the banner text, topic headers, and dates to the workbook, and applies formatting such as font styles and cell merging. Finally, it applies a border to all cells and saves the workbook as "format.xlsx".

Note: Make sure to install the openpyxl library before running this script.
'''

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active
'''current active worksheet is the first sheet in the workbook'''

# Banner Font
banner_font = Font(size=26, bold=True, italic=True)

# Topic Font
topic_font = Font(size=16, bold=True, italic=True)

file = 'format.xlsx'
'''file name to save the workbook as format.xlsx'''

def generate_format(month_days:list,year:str='2024',month:str='march') -> None:
    '''
    This function generates a formatted Excel workbook ```starting-template``` for a weekly roaster.
    '''
    # Write the banner text to the first row and format it
    ws['A1'] = 'Weekly Roaster ' + month.capitalize() + ' ' + year
    ws['A1'].font = banner_font


    # Write "Date" to the first cell of the second row
    ws.merge_cells('A2:A3')  # This will merge cells B2 and B3
    ws['A2'] = 'SL No'  # Now B2 refers to the merged cell
    ws['A2'].font = topic_font
    ws['A2'].alignment = Alignment(horizontal='center',vertical='center')  # Center the text in the cell

    ws.merge_cells('B2:B3')  # This will merge cells C2 and C3
    ws['B2'] = 'Operator and Helper Name'  # Now C2 refers to the merged cell
    ws['B2'].font = topic_font
    ws['B2'].alignment = Alignment(horizontal='center',vertical='center')  # Center the text in the cell

    ws.merge_cells('C2:C3')  # This will merge cells C2 and C3
    ws['C2'] = 'Designation'
    ws['C2'].font = topic_font
    ws['C2'].alignment = Alignment(horizontal='center',vertical='center')  # Center the text in the cell


    ws['D2'] = 'Date'
    ws['D2'].font = topic_font
    ws['D3'] = 'Day'
    ws['D3'].font = topic_font

    # Write the days of the week to the cells of the third row

    # label 1 upto 31 
    for i in range(len(month_days)):
        date_cell = ws.cell(row=2, column=i+5, value=i+1,)  # Write the day to the cell
        date_cell.alignment = Alignment(horizontal='center')
        date_cell.font = Font(bold=True)
        _ = ws.cell(row=3, column=i+5, value=month_days[i])  # Write the day to the cell
        _.alignment = Alignment(horizontal='center')


    # Create a Side object with desired style
    global border_all
    
    def border_all() -> None:
        '''
        - This function applies a border to all cells in the workbook.
        '''
        side = Side(border_style='thin', color='000000')  # '000000' is black

        # Create a Border object using the Side object
        border = Border(top=side, right=side, bottom=side, left=side)

        # Apply the border to the cells
        for row in ws.iter_rows(min_row=2): # skip the first row
            for cell in row:
                cell.border = border
    border_all()

# 
# # incert a new row
# def incert_row() -> None:
#     '''
#     This function incerts a new row in the workbook.
#     '''
#     # Incert a new row
#     ws.insert_rows(1)

# # write
# def write(row:int, column:int, value:str) -> None:
#     '''
#     This function writes a value to the specified cell in the workbook.
#     '''
#     ws.cell(row=row, column=column, value=value)



# Save the workbook | end
def save_workbook(file:str=file) -> None:
    # Save the workbook
    wb.save(file)

#  row 4 use-only
data_font = Font(size=11, bold=True, italic=False)
# SL No | A
def add_sl_no(index_:int=4, payload:int=1):
    '''
    - index_ : int : index of the row {```must be greater than =4```}
    '''
    A = 'A'+str(index_)
    ws[A] = payload
    ws[A].font = data_font
    ws[A].alignment = Alignment(horizontal='center',vertical='center')
    # ws[A].border = border

# Name | B
def add_name(payload:str='Operator_name',index_:int=4):
    '''
    - index_ : int : index of the row {```must be greater than =4```}
    '''
    # index of 
    B = 'B'+str(index_)
    ws[B] = payload 
    ws[B].font = data_font
    ws[B].alignment = Alignment(wrap_text=True,vertical='center')
    
    # ws[B].border = border

# Designation | C
def add_designation(payload:str='Highly Skilled',index_:int=4):
    '''
    - index_ : int : index of the row {```must be greater than =4```}
    '''
    # index of 
    C = 'C'+str(index_)
    ws[C] = payload
    ws[C].font = data_font
    ws[C].alignment = Alignment(horizontal='center',vertical='center')
    # ws[C].border = border

def add_data(manpower:list=any,designation:list=any) -> None:
    '''
    - manpower : list : list of employees
    - designation : list : list of designations
    #
    ```adds data to the workbook.```
    '''
    _ = 4
    for employee,desig in zip(manpower,designation):
        add_sl_no(index_=_,payload=manpower.index(employee)+1)
        # name_ = employee + '\n'
        add_name(employee,index_=_)
        add_designation(desig,index_=_)
        print(_)
        _ += 1
    
    border_all()

# #  Example
# manpower = ['john doe'+'\n'+'9547709029','jim doe'+ '\n' + '95058458880','Mak ane']
# designation = ['Plumber','Electrician','Carpenter']

# generate_format(month_days=['Mon','Tue','Wed','Thu','Fri','Sat','Sun'])
# add_data(manpower=manpower,designation=designation)

# # saving the workbook
# save_workbook(file=file)

#  Example end



